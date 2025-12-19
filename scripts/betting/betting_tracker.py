#!/usr/bin/env python3
"""Betting Tracker for NCAA CBB - Track picks, open bets, and daily results.

This script provides a complete betting workflow:
1. Generate betting slips (Excel) from KenPom analysis
2. Track open bets from overtime.ag
3. Track daily figures (P&L) from overtime.ag
4. Analyze end-of-day results

Usage:
    # Generate betting slip for today
    uv run python scripts/betting/betting_tracker.py slip

    # Track open bets
    uv run python scripts/betting/betting_tracker.py open-bets

    # Get daily figures
    uv run python scripts/betting/betting_tracker.py daily-figures

    # Analyze results (run at end of day)
    uv run python scripts/betting/betting_tracker.py analyze --date 2025-12-18

    # Full workflow
    uv run python scripts/betting/betting_tracker.py all
"""

import argparse
import asyncio
import json
import sys
from datetime import datetime
from pathlib import Path

import httpx

# Add src to path
sys.path.insert(0, str(Path(__file__).parent.parent.parent / "src"))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from playwright.async_api import async_playwright

from kenp0m_sp0rts_analyzer.api_client import KenPomAPI
from kenp0m_sp0rts_analyzer.luck_regression import LuckRegressionAnalyzer
from kenp0m_sp0rts_analyzer.utils import (
    normalize_team_name,
)

# ============================================================================
# BETTING SLIP GENERATION
# ============================================================================


def generate_betting_slip(
    date: str | None = None,
    min_edge: float = 3.0,
    output_dir: Path | None = None,
) -> Path:
    """Generate Excel betting slip from KenPom analysis.

    Args:
        date: Game date in YYYY-MM-DD format (default: today)
        min_edge: Minimum edge to include in slip
        output_dir: Output directory (default: reports/bet_slips/)

    Returns:
        Path to generated Excel file.
    """
    if date is None:
        date = datetime.now().strftime("%Y-%m-%d")

    if output_dir is None:
        output_dir = (
            Path(__file__).parent.parent.parent / "reports" / "bet_slips"
        )

    output_dir.mkdir(parents=True, exist_ok=True)

    # Format date for filename (YYYYMMDD)
    date_suffix = date.replace("-", "")
    output_path = output_dir / f"NCAA CBB Betting Slip {date_suffix}.xlsx"

    print(f"[INFO] Generating betting slip for {date}...")

    # Load Vegas lines
    vegas_path = (
        Path(__file__).parent.parent.parent
        / "data"
        / "overtime_cbb_lines.json"
    )
    if not vegas_path.exists():
        print(
            "[ERROR] No Vegas lines found. Run scrape_overtime_cbb.py first."
        )
        return None

    with vegas_path.open() as f:
        vegas_data = json.load(f)

    # Build Vegas lookup
    vegas_lines = {}
    for g in vegas_data["games"]:
        away = normalize_team_name(g["away_team"])
        home = normalize_team_name(g["home_team"])
        vegas_lines[(away, home)] = {
            "spread": g["spread"],
            "total": g["total"],
            "away_ml": g["away_ml"],
            "home_ml": g["home_ml"],
            "away_team": g["away_team"],
            "home_team": g["home_team"],
        }

    # Get KenPom predictions
    api = KenPomAPI()
    fanmatch = api.get_fanmatch(date)

    if not fanmatch.data:
        print(f"[ERROR] No games found for {date}")
        return None

    # Load ratings for luck analysis
    luck_analyzer = LuckRegressionAnalyzer()
    ratings_cache: dict[str, dict] = {}
    try:
        ratings = api.get_ratings(year=int(date[:4]))
        for team in ratings.data:
            team_name = team.get("TeamName", "")
            if team_name:
                ratings_cache[team_name] = team
        print(
            f"[OK] Loaded {len(ratings_cache)} team ratings for luck analysis"
        )
    except Exception as e:
        print(f"[WARN] Could not load ratings for luck: {e}")

    # Analyze games and find value picks
    spread_picks = []
    total_picks = []

    for game in fanmatch.data:
        visitor = game.get("Visitor", "")
        home = game.get("Home", "")

        visitor_pred = game["VisitorPred"]
        home_pred = game["HomePred"]
        home_wp = game["HomeWP"]

        # Calculate KenPom spread (positive = home underdog)
        kenpom_spread = home_pred - visitor_pred
        kp_spread = -kenpom_spread  # Convert to spread notation
        kenpom_total = visitor_pred + home_pred

        # Look up Vegas
        norm_away = normalize_team_name(visitor)
        norm_home = normalize_team_name(home)
        vegas = vegas_lines.get((norm_away, norm_home), {})

        if not vegas:
            continue

        v_spread = vegas.get("spread")
        v_total = vegas.get("total")

        # Get luck values for luck regression analysis
        visitor_luck = None
        home_luck = None
        luck_edge = 0.0

        # Look up in ratings cache
        for team_data in ratings_cache.values():
            team_name = team_data.get("TeamName", "")
            if team_name == visitor:
                visitor_luck = team_data.get("Luck", 0.0)
            if team_name == home:
                home_luck = team_data.get("Luck", 0.0)

        # Try normalized lookup if direct match failed
        if visitor_luck is None or home_luck is None:
            for team_data in ratings_cache.values():
                team_name = team_data.get("TeamName", "")
                norm_name = normalize_team_name(team_name)
                if norm_name == norm_away and visitor_luck is None:
                    visitor_luck = team_data.get("Luck", 0.0)
                if norm_name == norm_home and home_luck is None:
                    home_luck = team_data.get("Luck", 0.0)

        # Calculate luck edge if we have both values
        if visitor_luck is not None and home_luck is not None:
            # Get AdjEM values
            visitor_em = ratings_cache.get(visitor, {}).get("AdjEM", 0)
            home_em = ratings_cache.get(home, {}).get("AdjEM", 0)

            luck_result = luck_analyzer.analyze_matchup_luck(
                team1_name=visitor,
                team1_adjEM=visitor_em,
                team1_luck=visitor_luck,
                team2_name=home,
                team2_adjEM=home_em,
                team2_luck=home_luck,
                games_remaining=15,
                neutral_site=False,
                home_court_advantage=3.5,
            )
            luck_edge = luck_result.luck_edge

        if v_spread is not None:
            spread_edge = kp_spread - v_spread
            composite_edge = spread_edge + luck_edge

            # Use COMPOSITE edge for value determination
            if abs(composite_edge) >= min_edge:
                # Determine pick based on composite
                if composite_edge < 0:
                    # Composite favors HOME - bet HOME
                    if v_spread > 0:
                        pick = f"{home} +{v_spread}"
                        pick_team = home
                    else:
                        pick = f"{home} {v_spread}"
                        pick_team = home
                else:
                    # Composite favors AWAY - bet AWAY
                    if v_spread > 0:
                        pick = f"{visitor} -{v_spread}"
                        pick_team = visitor
                    else:
                        pick = f"{visitor} +{abs(v_spread)}"
                        pick_team = visitor

                spread_picks.append(
                    {
                        "matchup": f"{visitor} @ {home}",
                        "pick": pick,
                        "pick_team": pick_team,
                        "kp_spread": kp_spread,
                        "vegas_spread": v_spread,
                        "kp_edge": spread_edge,
                        "luck_edge": luck_edge,
                        "composite_edge": composite_edge,
                        "visitor_luck": visitor_luck,
                        "home_luck": home_luck,
                        "home_wp": home_wp,
                        "bet_type": "SPREAD",
                    }
                )

        if v_total is not None:
            total_edge = kenpom_total - v_total

            if abs(total_edge) >= 5:
                if total_edge > 0:
                    pick = f"OVER {v_total}"
                else:
                    pick = f"UNDER {v_total}"

                total_picks.append(
                    {
                        "matchup": f"{visitor} @ {home}",
                        "pick": pick,
                        "kp_total": kenpom_total,
                        "vegas_total": v_total,
                        "edge": total_edge,
                        "bet_type": "TOTAL",
                    }
                )

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Betting Slip"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    value_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Title
    ws.merge_cells("A1:J1")
    ws["A1"] = f"NCAA CBB BETTING SLIP - {date}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Generated timestamp
    ws.merge_cells("A2:J2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].alignment = Alignment(horizontal="center")

    # Spread Picks Section
    row = 4
    ws.merge_cells(f"A{row}:J{row}")
    ws[f"A{row}"] = "SPREAD VALUE PICKS (w/ Luck Regression)"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    ws[f"A{row}"].fill = PatternFill(
        start_color="FFC000", end_color="FFC000", fill_type="solid"
    )

    row += 1
    headers = [
        "Matchup",
        "Pick",
        "KP Spread",
        "Vegas",
        "KP Edge",
        "Luck Edge",
        "COMPOSITE",
        "WP%",
        "Placed",
        "Result",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    row += 1
    for pick in sorted(
        spread_picks, key=lambda x: abs(x["composite_edge"]), reverse=True
    ):
        ws.cell(row=row, column=1, value=pick["matchup"]).border = thin_border
        ws.cell(row=row, column=2, value=pick["pick"]).border = thin_border
        ws.cell(row=row, column=2).fill = value_fill
        ws.cell(
            row=row, column=3, value=f"{pick['kp_spread']:+.1f}"
        ).border = thin_border
        ws.cell(
            row=row, column=4, value=f"{pick['vegas_spread']:+.1f}"
        ).border = thin_border
        ws.cell(
            row=row, column=5, value=f"{pick['kp_edge']:+.1f}"
        ).border = thin_border
        ws.cell(
            row=row, column=6, value=f"{pick['luck_edge']:+.1f}"
        ).border = thin_border
        # Highlight composite edge
        comp_cell = ws.cell(
            row=row, column=7, value=f"{pick['composite_edge']:+.1f}"
        )
        comp_cell.border = thin_border
        comp_cell.font = Font(bold=True)
        ws.cell(
            row=row, column=8, value=f"{pick['home_wp']}%"
        ).border = thin_border
        ws.cell(row=row, column=9, value="").border = thin_border  # Placed
        ws.cell(row=row, column=10, value="").border = thin_border  # Result
        row += 1

    if not spread_picks:
        ws.cell(row=row, column=1, value="No spread value picks today")
        row += 1

    # Total Picks Section
    row += 2
    ws.merge_cells(f"A{row}:J{row}")
    ws[f"A{row}"] = "TOTAL VALUE PICKS"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    ws[f"A{row}"].fill = PatternFill(
        start_color="70AD47", end_color="70AD47", fill_type="solid"
    )

    row += 1
    headers = [
        "Matchup",
        "Pick",
        "KP Total",
        "Vegas",
        "Edge",
        "",
        "",
        "",
        "Placed",
        "Result",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    row += 1
    for pick in sorted(
        total_picks, key=lambda x: abs(x["edge"]), reverse=True
    ):
        ws.cell(row=row, column=1, value=pick["matchup"]).border = thin_border
        ws.cell(row=row, column=2, value=pick["pick"]).border = thin_border
        ws.cell(row=row, column=2).fill = value_fill
        ws.cell(
            row=row, column=3, value=f"{pick['kp_total']:.1f}"
        ).border = thin_border
        ws.cell(
            row=row, column=4, value=f"{pick['vegas_total']}"
        ).border = thin_border
        # Bold the edge for totals
        edge_cell = ws.cell(row=row, column=5, value=f"{pick['edge']:+.1f}")
        edge_cell.border = thin_border
        edge_cell.font = Font(bold=True)
        ws.cell(row=row, column=6, value="").border = thin_border
        ws.cell(row=row, column=7, value="").border = thin_border
        ws.cell(row=row, column=8, value="").border = thin_border
        ws.cell(row=row, column=9, value="").border = thin_border  # Placed
        ws.cell(row=row, column=10, value="").border = thin_border  # Result
        row += 1

    if not total_picks:
        ws.cell(row=row, column=1, value="No total value picks today")
        row += 1

    # Summary Section
    row += 2
    ws[f"A{row}"] = "SUMMARY"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1
    ws[f"A{row}"] = f"Spread Picks: {len(spread_picks)}"
    row += 1
    ws[f"A{row}"] = f"Total Picks: {len(total_picks)}"
    row += 1
    ws[f"A{row}"] = f"Min Composite Edge: {min_edge} pts"
    row += 1
    ws[f"A{row}"] = "Note: Composite = KenPom Edge + Luck Regression Edge"

    # Adjust column widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 9
    ws.column_dimensions["F"].width = 9
    ws.column_dimensions["G"].width = 11
    ws.column_dimensions["H"].width = 6
    ws.column_dimensions["I"].width = 8
    ws.column_dimensions["J"].width = 8

    # Save
    wb.save(output_path)
    print(f"[OK] Betting slip saved: {output_path}")

    return output_path


# ============================================================================
# OPEN BETS TRACKING
# ============================================================================


def get_user_data_dir() -> Path:
    """Get persistent browser data directory for overtime.ag session."""
    data_dir = (
        Path.home() / ".cache" / "kenp0m_sp0rts_analyzer" / "overtime_browser"
    )
    data_dir.mkdir(parents=True, exist_ok=True)
    return data_dir


async def fetch_open_bets(headless: bool = False) -> dict:
    """Fetch open bets from overtime.ag.

    Uses a persistent browser context to maintain login session.
    First run requires manual login in the browser window.

    Args:
        headless: Run in headless mode. Use False for first login.

    Returns:
        Dictionary with open bets data.
    """
    print("[INFO] Fetching open bets from overtime.ag...")
    print("[INFO] Using persistent browser session.")

    if not headless:
        print("[INFO] Browser window will open. Log in if prompted.")

    user_data_dir = get_user_data_dir()

    async with async_playwright() as p:
        # Use persistent context to preserve login cookies
        context = await p.chromium.launch_persistent_context(
            user_data_dir=str(user_data_dir),
            headless=headless,
            viewport={"width": 1280, "height": 800},
        )
        page = context.pages[0] if context.pages else await context.new_page()

        try:
            # Navigate to sports section first
            await page.goto("https://overtime.ag/sports", timeout=30000)
            await asyncio.sleep(3)

            # Navigate to open bets page
            await page.goto(
                "https://overtime.ag/sports#/openBets", timeout=30000
            )
            await asyncio.sleep(3)

            # Try to get open bets via page content scraping
            result = await page.evaluate(
                """
                () => {
                    // Look for bet information on the page
                    const pageText = document.body.innerText;

                    // Check for "No open bets" message
                    if (pageText.includes('No open bets') ||
                        pageText.includes('no open wagers')) {
                        return {status: 'empty', message: 'No open bets found'};
                    }

                    // Check for login prompt
                    if (pageText.includes('Sign In') ||
                        pageText.includes('Log In') ||
                        pageText.includes('Please login')) {
                        return {status: 'not_logged_in', message: 'Login required'};
                    }

                    // Try to find bet elements
                    const betElements = document.querySelectorAll(
                        '[class*="bet"], [class*="wager"], [class*="open"]'
                    );
                    const bets = [];
                    betElements.forEach(el => {
                        const text = el.innerText.trim();
                        if (text && text.length > 10) {
                            bets.push(text.substring(0, 200));
                        }
                    });

                    if (bets.length > 0) {
                        return {status: 'found', bets: bets};
                    }

                    // Return page content summary
                    return {
                        status: 'unknown',
                        content: pageText.substring(0, 1000)
                    };
                }
            """
            )

            await context.close()
            return result

        except Exception as e:
            await context.close()
            return {"error": str(e)}


def display_open_bets(data: dict) -> None:
    """Display open bets in a formatted table."""
    if "error" in data:
        print(f"[ERROR] {data['error']}")
        return

    print("\n" + "=" * 80)
    print("OPEN BETS")
    print("=" * 80)

    status = data.get("status", "unknown")

    if status == "empty":
        print("\n[INFO] No open bets found")
        return

    if status == "not_logged_in":
        print("\n[WARN] Not logged in to overtime.ag")
        print("       Run with visible browser to log in:")
        print(
            "       uv run python scripts/betting/betting_tracker.py open-bets"
        )
        return

    if status == "found":
        bets = data.get("bets", [])
        print(f"\n[INFO] Found {len(bets)} bet entries:")
        for i, bet in enumerate(bets, 1):
            print(f"\n  [{i}] {bet}")
        return

    # Legacy API format
    if "d" in data and data["d"]:
        bets = (
            data["d"].get("Data", [])
            if isinstance(data["d"], dict)
            else data["d"]
        )
        for bet in bets:
            print(f"\nBet ID: {bet.get('BetId', 'N/A')}")
            print(f"  Type: {bet.get('BetType', 'N/A')}")
            print(f"  Risk: ${bet.get('RiskAmount', 0):.2f}")
            print(f"  Win: ${bet.get('WinAmount', 0):.2f}")
            print(f"  Status: {bet.get('Status', 'N/A')}")
            legs = bet.get("Legs", [])
            for leg in legs:
                print(f"    -> {leg.get('Description', 'N/A')}")
        return

    # Unknown format - show raw content
    content = data.get("content", "No content available")
    print(f"\n[INFO] Page content:\n{content[:500]}...")


# ============================================================================
# DAILY FIGURES TRACKING
# ============================================================================


async def fetch_daily_figures(
    date: str | None = None, headless: bool = False
) -> dict:
    """Fetch daily P&L figures from overtime.ag.

    Uses a persistent browser context to maintain login session.

    Args:
        date: Date in YYYY-MM-DD format (default: today)
        headless: Run in headless mode. Use False for first login.

    Returns:
        Dictionary with daily figures data.
    """
    if date is None:
        date = datetime.now().strftime("%Y-%m-%d")

    print(f"[INFO] Fetching daily figures for {date}...")
    print("[INFO] Using persistent browser session.")

    user_data_dir = get_user_data_dir()

    async with async_playwright() as p:
        context = await p.chromium.launch_persistent_context(
            user_data_dir=str(user_data_dir),
            headless=headless,
            viewport={"width": 1280, "height": 800},
        )
        page = context.pages[0] if context.pages else await context.new_page()

        try:
            # Navigate to daily figures page
            await page.goto(
                "https://overtime.ag/sports#/dailyFigures", timeout=30000
            )
            await asyncio.sleep(3)

            # Try to scrape daily figures from page
            result = await page.evaluate(
                """
                () => {
                    const pageText = document.body.innerText;

                    // Check for login prompt
                    if (pageText.includes('Sign In') ||
                        pageText.includes('Please login')) {
                        return {status: 'not_logged_in'};
                    }

                    // Look for P&L data patterns
                    const figures = {};

                    // Try to find balance/P&L elements
                    const balanceEl = document.querySelector('[class*="balance"]');
                    if (balanceEl) {
                        figures.balance = balanceEl.innerText;
                    }

                    // Look for daily summary table
                    const tables = document.querySelectorAll('table');
                    const tableData = [];
                    tables.forEach(table => {
                        tableData.push(table.innerText.substring(0, 500));
                    });

                    if (tableData.length > 0) {
                        figures.tables = tableData;
                    }

                    // Get page content summary
                    figures.content = pageText.substring(0, 1500);
                    figures.status = 'scraped';

                    return figures;
                }
            """
            )

            await context.close()
            return result

        except Exception as e:
            await context.close()
            return {"error": str(e)}


def display_daily_figures(data: dict, date: str) -> None:
    """Display daily figures."""
    if "error" in data:
        print(f"[ERROR] {data['error']}")
        return

    print("\n" + "=" * 80)
    print(f"DAILY FIGURES - {date}")
    print("=" * 80)

    status = data.get("status", "unknown")

    if status == "not_logged_in":
        print("\n[WARN] Not logged in to overtime.ag")
        print("       Run with visible browser to log in first.")
        return

    if status == "scraped":
        if data.get("balance"):
            print(f"\nBalance: {data['balance']}")

        if data.get("tables"):
            print("\nData Tables Found:")
            for i, table in enumerate(data["tables"], 1):
                print(f"\n  Table {i}:")
                print(f"  {table[:300]}...")

        if data.get("content"):
            print("\nPage Summary:")
            content = data["content"]
            # Look for P&L patterns in content
            lines = content.split("\n")
            for line in lines:
                line = line.strip()
                keywords = ["p&l", "profit", "loss", "win", "total"]
                has_keyword = any(kw in line.lower() for kw in keywords)
                if has_keyword and len(line) < 100:
                    print(f"  {line}")
        return

    # Legacy API format
    if "d" in data and data["d"]:
        figures = data["d"]
        if isinstance(figures, dict):
            print(f"\nNet P&L: ${figures.get('NetPL', 0):.2f}")
            print(f"Total Risked: ${figures.get('TotalRisked', 0):.2f}")
            print(f"Total Won: ${figures.get('TotalWon', 0):.2f}")
            print(f"Bets Placed: {figures.get('BetsPlaced', 0)}")
            print(f"Wins: {figures.get('Wins', 0)}")
            print(f"Losses: {figures.get('Losses', 0)}")
        else:
            print(f"\nRaw data: {figures}")
        return

    print("[INFO] No daily figures available")


# ============================================================================
# END-OF-DAY ANALYSIS
# ============================================================================


def analyze_results(date: str | None = None) -> dict:
    """Analyze betting results for a given date.

    Compares KenPom predictions with actual results.

    Args:
        date: Date in YYYY-MM-DD format (default: today)

    Returns:
        Dictionary with analysis results.
    """
    if date is None:
        date = datetime.now().strftime("%Y-%m-%d")

    date_suffix = date.replace("-", "")

    print(f"\n[INFO] Analyzing results for {date}...")

    # Load betting slip
    slip_path = (
        Path(__file__).parent.parent.parent
        / "reports"
        / "bet_slips"
        / f"NCAA CBB Betting Slip {date_suffix}.xlsx"
    )

    if not slip_path.exists():
        print(f"[WARN] No betting slip found for {date}")
        print(f"       Expected: {slip_path}")
        return {"error": "No betting slip found"}

    # Load results from tracking database (if available)
    results_path = (
        Path(__file__).parent.parent.parent
        / "data"
        / "results"
        / f"results_{date_suffix}.json"
    )

    results = {
        "date": date,
        "slip_path": str(slip_path),
        "picks": [],
        "summary": {},
    }

    if results_path.exists():
        with results_path.open() as f:
            actual_results = json.load(f)

        # Analyze each pick
        wins = 0
        losses = 0
        pushes = 0

        for pick in actual_results.get("picks", []):
            if pick.get("result") == "WIN":
                wins += 1
            elif pick.get("result") == "LOSS":
                losses += 1
            else:
                pushes += 1

            results["picks"].append(pick)

        total = wins + losses
        win_pct = (wins / total * 100) if total > 0 else 0

        results["summary"] = {
            "wins": wins,
            "losses": losses,
            "pushes": pushes,
            "total": total,
            "win_percentage": win_pct,
        }

        print("\n" + "=" * 80)
        print(f"RESULTS ANALYSIS - {date}")
        print("=" * 80)
        print(f"\nRecord: {wins}-{losses}-{pushes}")
        print(f"Win %: {win_pct:.1f}%")

        print("\nPick-by-Pick:")
        for pick in results["picks"]:
            status = pick.get("result", "PENDING")
            emoji = (
                "[WIN]"
                if status == "WIN"
                else "[LOSS]"
                if status == "LOSS"
                else "[--]"
            )
            print(f"  {emoji} {pick.get('pick', 'N/A')}")
            if pick.get("notes"):
                print(f"       Notes: {pick['notes']}")

    else:
        print("[INFO] No results file found. Create one at:")
        print(f"       {results_path}")
        print("\nExpected format:")
        print(
            """
{
  "date": "2025-12-18",
  "picks": [
    {"pick": "Indiana St. +5", "result": "WIN", "final_margin": 3},
    {"pick": "Temple/Davidson OVER 144", "result": "LOSS", "final_total": 138}
  ]
}
"""
        )

    return results


# ============================================================================
# LIVE SCORE TRACKING
# ============================================================================

ESPN_SCOREBOARD_URL = (
    "https://site.api.espn.com/apis/site/v2/sports/basketball/mens-college-basketball"
    "/scoreboard"
)

# Common college basketball mascots to strip from ESPN names
ESPN_MASCOTS = [
    "Bulldogs", "Tigers", "Wildcats", "Catamounts", "Eagles", "Bears",
    "Cavaliers", "Tar Heels", "Blue Devils", "Cardinals", "Wolfpack",
    "Hoosiers", "Boilermakers", "Hawkeyes", "Spartans", "Wolverines",
    "Buckeyes", "Nittany Lions", "Fighting Irish", "Seminoles", "Hurricanes",
    "Yellow Jackets", "Crimson Tide", "Volunteers", "Gators", "Gamecocks",
    "Razorbacks", "Longhorns", "Jayhawks", "Cyclones", "Sooners", "Cowboys",
    "Red Raiders", "Horned Frogs", "Cougars", "Mountaineers", "Panthers",
    "Orange", "Hokies", "Demon Deacons", "Commodores", "Rebels", "Aggies",
    "Sun Devils", "Bruins", "Trojans", "Ducks", "Beavers", "Huskies",
    "Golden Bears", "Buffaloes", "Utes", "Aztecs", "Toreros", "Broncos",
    "Gaels", "Pilots", "Zags", "Gonzaga Bulldogs", "Wave", "Saints",
    "Bearcats", "Musketeers", "Flyers", "Explorers", "Hawks", "Owls",
    "Minutemen", "Friars", "Bluejays", "Red Storm", "Hoyas", "Pirates",
    "Golden Eagles", "Peacocks", "Bonnies", "Rams", "Billikens", "Colonials",
    "Dukes", "Spiders", "Monarchs", "Phoenix", "49ers", "Chanticleers",
    "Paladins", "Terriers", "Keydets", "Tribe", "Pride", "Dolphins",
    "Flames", "Scarlet Knights", "Sycamores", "Penguins", "Redhawks",
    "Bobcats", "Thundering Herd", "Golden Flashes", "Chippewas", "Broncos",
    "Rockets", "Falcons", "Zips", "RedHawks", "Bulls", "Leopards",
    "Mountain Hawks", "Bison", "Crusaders", "Mastodons", "Leathernecks",
    "Fighting Illini", "Badgers", "Golden Gophers", "Cornhuskers", "Wildcats",
    "Terrapins", "Blue Hens", "Retrievers", "Great Danes", "Seawolves",
    "River Hawks", "Catamounts", "Black Bears", "Wildcats", "Huskies",
]


def strip_espn_mascot(team_name: str) -> str:
    """Strip ESPN mascot from team name to get just the school name.

    ESPN returns names like "Georgia Bulldogs" but we need "Georgia".

    Args:
        team_name: Full ESPN team name with mascot.

    Returns:
        School name without mascot.
    """
    name = team_name.strip()

    # Sort mascots by length (longest first) to avoid partial matches
    for mascot in sorted(ESPN_MASCOTS, key=len, reverse=True):
        if name.endswith(f" {mascot}"):
            return name[: -(len(mascot) + 1)].strip()

    # No mascot found, return as-is
    return name


async def fetch_live_scores(date: str | None = None) -> list[dict]:
    """Fetch live scores from ESPN API.

    Args:
        date: Date in YYYY-MM-DD format (default: today)

    Returns:
        List of game dictionaries with scores and status.
    """
    if date is None:
        date = datetime.now().strftime("%Y-%m-%d")

    # Format date for ESPN API (YYYYMMDD)
    espn_date = date.replace("-", "")

    url = f"{ESPN_SCOREBOARD_URL}?dates={espn_date}"

    games = []
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.get(url)
            response.raise_for_status()
            data = response.json()

        for event in data.get("events", []):
            competition = event.get("competitions", [{}])[0]
            competitors = competition.get("competitors", [])

            if len(competitors) != 2:
                continue

            # ESPN has home team first, away team second
            home_data = None
            away_data = None
            for comp in competitors:
                if comp.get("homeAway") == "home":
                    home_data = comp
                else:
                    away_data = comp

            if not home_data or not away_data:
                continue

            status = event.get("status", {})
            status_type = status.get("type", {}).get(
                "name", "STATUS_SCHEDULED"
            )
            period = status.get("period", 0)
            clock = status.get("displayClock", "0:00")

            # Determine game state
            if status_type == "STATUS_FINAL":
                game_status = "FINAL"
            elif status_type == "STATUS_IN_PROGRESS":
                if period == 1:
                    game_status = f"1H {clock}"
                elif period == 2:
                    game_status = f"2H {clock}"
                else:
                    game_status = (
                        f"OT{period - 2} {clock}" if period > 2 else "LIVE"
                    )
            elif status_type == "STATUS_HALFTIME":
                game_status = "HALF"
            else:
                game_status = "PRE"

            home_score = int(home_data.get("score", 0) or 0)
            away_score = int(away_data.get("score", 0) or 0)

            games.append(
                {
                    "home_team": home_data.get("team", {}).get(
                        "displayName", "Unknown"
                    ),
                    "away_team": away_data.get("team", {}).get(
                        "displayName", "Unknown"
                    ),
                    "home_abbrev": home_data.get("team", {}).get(
                        "abbreviation", "???"
                    ),
                    "away_abbrev": away_data.get("team", {}).get(
                        "abbreviation", "???"
                    ),
                    "home_score": home_score,
                    "away_score": away_score,
                    "total": home_score + away_score,
                    "margin": home_score
                    - away_score,  # Positive = home winning
                    "status": game_status,
                    "period": period,
                    "clock": clock,
                    "is_final": status_type == "STATUS_FINAL",
                    "is_live": status_type == "STATUS_IN_PROGRESS",
                }
            )

    except Exception as e:
        print(f"[ERROR] Failed to fetch scores: {e}")

    return games


def calculate_cover_status(
    pick: dict,
    game: dict,
) -> dict:
    """Calculate whether a pick is covering, pushing, or not covering.

    Args:
        pick: Pick dictionary from betting slip
        game: Live game data from ESPN

    Returns:
        Dictionary with cover status details.
    """
    bet_type = pick.get("bet_type", "SPREAD")
    status = game.get("status", "PRE")
    is_final = game.get("is_final", False)
    is_live = game.get("is_live", False)

    result = {
        "pick": pick.get("pick", "N/A"),
        "matchup": pick.get("matchup", "N/A"),
        "status": status,
        "home_score": game.get("home_score", 0),
        "away_score": game.get("away_score", 0),
        "is_final": is_final,
        "cover_status": "PENDING",
        "margin_vs_spread": None,
        "current_total": game.get("total", 0),
    }

    if not is_live and not is_final:
        result["cover_status"] = "NOT STARTED"
        return result

    if bet_type == "SPREAD":
        vegas_spread = pick.get("vegas_spread", 0)
        pick_str = pick.get("pick", "")

        # Determine which team we're betting on
        home_team = (
            pick.get("matchup", "").split(" @ ")[-1]
            if " @ " in pick.get("matchup", "")
            else ""
        )
        away_team = (
            pick.get("matchup", "").split(" @ ")[0]
            if " @ " in pick.get("matchup", "")
            else ""
        )

        current_margin = game.get("margin", 0)  # Positive = home winning

        # Check if we bet on home or away
        if home_team and home_team in pick_str:
            # We bet on home team
            # Home team covers if: current_margin > -spread
            # e.g., Home +5: covers if loses by less than 5
            spread_for_home = vegas_spread
            margin_vs_spread = current_margin - (-spread_for_home)
            if margin_vs_spread > 0:
                result["cover_status"] = "COVERING" if is_live else "WIN"
            elif margin_vs_spread < 0:
                result["cover_status"] = "NOT COVERING" if is_live else "LOSS"
            else:
                result["cover_status"] = "PUSH"
            result["margin_vs_spread"] = margin_vs_spread
        elif away_team and away_team in pick_str:
            # We bet on away team
            # Away covers if: -current_margin > -spread (for away)
            spread_for_away = -vegas_spread
            away_margin = -current_margin  # Positive = away winning
            margin_vs_spread = away_margin - (-spread_for_away)
            if margin_vs_spread > 0:
                result["cover_status"] = "COVERING" if is_live else "WIN"
            elif margin_vs_spread < 0:
                result["cover_status"] = "NOT COVERING" if is_live else "LOSS"
            else:
                result["cover_status"] = "PUSH"
            result["margin_vs_spread"] = margin_vs_spread

    elif bet_type == "TOTAL":
        vegas_total = pick.get("vegas_total", 0)
        current_total = game.get("total", 0)
        pick_str = pick.get("pick", "")

        is_over = "OVER" in pick_str.upper()

        if is_over:
            margin_vs_total = current_total - vegas_total
            if current_total > vegas_total:
                result["cover_status"] = "COVERING" if is_live else "WIN"
            elif current_total < vegas_total:
                result["cover_status"] = "NOT COVERING" if is_live else "LOSS"
            else:
                result["cover_status"] = "PUSH"
            result["margin_vs_spread"] = margin_vs_total
        else:  # UNDER
            margin_vs_total = vegas_total - current_total
            if current_total < vegas_total:
                result["cover_status"] = "COVERING" if is_live else "WIN"
            elif current_total > vegas_total:
                result["cover_status"] = "NOT COVERING" if is_live else "LOSS"
            else:
                result["cover_status"] = "PUSH"
            result["margin_vs_spread"] = margin_vs_total

    return result


async def track_picks_live(date: str | None = None) -> None:
    """Track betting picks in real-time.

    Args:
        date: Date in YYYY-MM-DD format (default: today)
    """
    if date is None:
        date = datetime.now().strftime("%Y-%m-%d")

    print(f"\n[INFO] Tracking live picks for {date}...")
    print("[INFO] Press Ctrl+C to stop tracking\n")

    # Load picks from Vegas lines
    vegas_path = (
        Path(__file__).parent.parent.parent / "data" / "vegas_lines.json"
    )

    if not vegas_path.exists():
        print("[ERROR] No Vegas lines found. Generate a betting slip first.")
        return

    with vegas_path.open() as f:
        vegas_data = json.load(f)

    # Get KenPom analysis for picks
    api = KenPomAPI()

    try:
        fanmatch = api.get_fanmatch(date)
    except Exception as e:
        print(f"[ERROR] Failed to get KenPom data: {e}")
        return

    # Build picks list based on edges
    picks = []

    # Load ratings for luck analysis
    luck_analyzer = LuckRegressionAnalyzer()
    ratings_cache: dict[str, dict] = {}
    try:
        ratings = api.get_ratings(year=int(date[:4]))
        for team in ratings.data:
            team_name = team.get("TeamName", "")
            if team_name:
                ratings_cache[team_name] = team
    except Exception:
        pass

    for game in fanmatch.data:
        visitor = game.get("Visitor", "")
        home = game.get("Home", "")

        visitor_pred = game["VisitorPred"]
        home_pred = game["HomePred"]

        kenpom_spread = home_pred - visitor_pred
        kp_spread = -kenpom_spread
        kenpom_total = visitor_pred + home_pred

        # Find Vegas line
        norm_away = normalize_team_name(visitor)
        norm_home = normalize_team_name(home)

        vegas_line = None
        for g in vegas_data.get("games", []):
            if (
                normalize_team_name(g["away_team"]) == norm_away
                and normalize_team_name(g["home_team"]) == norm_home
            ):
                vegas_line = g
                break

        if not vegas_line:
            continue

        v_spread = vegas_line.get("spread")
        v_total = vegas_line.get("total")

        # Calculate luck edge
        luck_edge = 0.0
        for team_data in ratings_cache.values():
            team_name = team_data.get("TeamName", "")
            if team_name == visitor:
                visitor_luck = team_data.get("Luck", 0.0)
            if team_name == home:
                home_luck = team_data.get("Luck", 0.0)

        if "visitor_luck" in dir() and "home_luck" in dir():
            try:
                luck_result = luck_analyzer.analyze_matchup_luck(
                    team1_name=visitor,
                    team1_adjEM=ratings_cache.get(visitor, {}).get("AdjEM", 0),
                    team1_luck=visitor_luck,
                    team2_name=home,
                    team2_adjEM=ratings_cache.get(home, {}).get("AdjEM", 0),
                    team2_luck=home_luck,
                    games_remaining=15,
                    neutral_site=False,
                    home_court_advantage=3.5,
                )
                luck_edge = luck_result.luck_edge
            except Exception:
                luck_edge = 0.0

        # Check spread edge
        if v_spread is not None:
            spread_edge = kp_spread - v_spread
            composite_edge = spread_edge + luck_edge

            if abs(composite_edge) >= 3:
                if composite_edge < 0:
                    if v_spread > 0:
                        pick_str = f"{home} +{v_spread}"
                    else:
                        pick_str = f"{home} {v_spread}"
                else:
                    if v_spread > 0:
                        pick_str = f"{visitor} -{v_spread}"
                    else:
                        pick_str = f"{visitor} +{abs(v_spread)}"

                picks.append(
                    {
                        "matchup": f"{visitor} @ {home}",
                        "pick": pick_str,
                        "bet_type": "SPREAD",
                        "vegas_spread": v_spread,
                        "composite_edge": composite_edge,
                    }
                )

        # Check total edge
        if v_total is not None:
            total_edge = kenpom_total - v_total

            if abs(total_edge) >= 5:
                if total_edge > 0:
                    pick_str = f"OVER {v_total}"
                else:
                    pick_str = f"UNDER {v_total}"

                picks.append(
                    {
                        "matchup": f"{visitor} @ {home}",
                        "pick": pick_str,
                        "bet_type": "TOTAL",
                        "vegas_total": v_total,
                        "edge": total_edge,
                    }
                )

    if not picks:
        print("[INFO] No value picks to track")
        return

    print(f"[OK] Tracking {len(picks)} picks\n")

    # Fetch live scores and display status
    scores = await fetch_live_scores(date)

    if not scores:
        print("[WARN] No live scores available")
        return

    # Build lookup by team names (strip ESPN mascots first)
    scores_by_team: dict[str, dict] = {}
    for game in scores:
        # ESPN returns "Georgia Bulldogs" -> strip to "Georgia" -> normalize
        home_stripped = strip_espn_mascot(game["home_team"])
        away_stripped = strip_espn_mascot(game["away_team"])
        home = normalize_team_name(home_stripped)
        away = normalize_team_name(away_stripped)
        scores_by_team[home] = game
        scores_by_team[away] = game
        # Also add with original stripped name for fallback matching
        scores_by_team[home_stripped] = game
        scores_by_team[away_stripped] = game

    # Display tracking table
    print("=" * 85)
    print("LIVE PICK TRACKER")
    print("=" * 85)
    print(f"{'PICK':<35} {'SCORE':<15} {'STATUS':<10} {'COVER':<15}")
    print("-" * 85)

    covering = 0
    not_covering = 0
    pending = 0

    for pick in picks:
        matchup = pick["matchup"]
        parts = matchup.split(" @ ")
        away_team = normalize_team_name(parts[0]) if len(parts) > 0 else ""
        home_team = normalize_team_name(parts[1]) if len(parts) > 1 else ""

        # Find game in scores
        game = scores_by_team.get(home_team) or scores_by_team.get(away_team)

        if not game:
            print(f"{pick['pick']:<35} {'N/A':<15} {'--':<10} {'NO DATA':<15}")
            pending += 1
            continue

        # Calculate cover status
        status = calculate_cover_status(pick, game)

        score_str = f"{game['away_score']}-{game['home_score']}"
        game_status = game["status"]

        cover = status["cover_status"]
        if cover == "COVERING" or cover == "WIN":
            cover_display = f"[OK] {cover}"
            covering += 1
        elif cover == "NOT COVERING" or cover == "LOSS":
            cover_display = f"[X] {cover}"
            not_covering += 1
        elif cover == "PUSH":
            cover_display = "[--] PUSH"
            pending += 1
        else:
            cover_display = f"[?] {cover}"
            pending += 1

        # Add margin info if available
        if status.get("margin_vs_spread") is not None:
            margin = status["margin_vs_spread"]
            cover_display += f" ({margin:+.0f})"

        row = f"{pick['pick']:<35} {score_str:<15} "
        row += f"{game_status:<10} {cover_display:<15}"
        print(row)

    print("-" * 85)
    summary = f"\nSUMMARY: Covering: {covering} | "
    summary += f"Not Covering: {not_covering} | Pending: {pending}"
    print(summary)
    print("=" * 85)


# ============================================================================
# MAIN CLI
# ============================================================================


def main():
    parser = argparse.ArgumentParser(
        description="NCAA CBB Betting Tracker",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )

    subparsers = parser.add_subparsers(dest="command", help="Commands")

    # Slip command
    slip_parser = subparsers.add_parser("slip", help="Generate betting slip")
    slip_parser.add_argument(
        "--date", "-d", help="Date (YYYY-MM-DD, default: today)"
    )
    slip_parser.add_argument(
        "--min-edge",
        type=float,
        default=3.0,
        help="Minimum edge (default: 3.0)",
    )

    # Open bets command
    subparsers.add_parser("open-bets", help="Fetch open bets from overtime.ag")

    # Daily figures command
    figures_parser = subparsers.add_parser(
        "daily-figures", help="Fetch daily P&L figures"
    )
    figures_parser.add_argument(
        "--date", "-d", help="Date (YYYY-MM-DD, default: today)"
    )

    # Analyze command
    analyze_parser = subparsers.add_parser("analyze", help="Analyze results")
    analyze_parser.add_argument(
        "--date", "-d", help="Date (YYYY-MM-DD, default: today)"
    )

    # All command
    all_parser = subparsers.add_parser("all", help="Run full workflow")
    all_parser.add_argument(
        "--date", "-d", help="Date (YYYY-MM-DD, default: today)"
    )

    # Live tracking command
    live_parser = subparsers.add_parser(
        "live", help="Track picks in real-time"
    )
    live_parser.add_argument(
        "--date", "-d", help="Date (YYYY-MM-DD, default: today)"
    )

    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        return 1

    print("=" * 80)
    print("NCAA CBB BETTING TRACKER")
    print("=" * 80)

    if args.command == "slip":
        generate_betting_slip(date=args.date, min_edge=args.min_edge)

    elif args.command == "open-bets":
        data = asyncio.run(fetch_open_bets())
        display_open_bets(data)

    elif args.command == "daily-figures":
        date = args.date or datetime.now().strftime("%Y-%m-%d")
        data = asyncio.run(fetch_daily_figures(date))
        display_daily_figures(data, date)

    elif args.command == "analyze":
        analyze_results(date=args.date)

    elif args.command == "all":
        date = args.date or datetime.now().strftime("%Y-%m-%d")
        print("\n[1/4] Generating betting slip...")
        generate_betting_slip(date=date)

        print("\n[2/4] Fetching open bets...")
        bets = asyncio.run(fetch_open_bets())
        display_open_bets(bets)

        print("\n[3/4] Fetching daily figures...")
        figures = asyncio.run(fetch_daily_figures(date))
        display_daily_figures(figures, date)

        print("\n[4/4] Analyzing results...")
        analyze_results(date=date)

    elif args.command == "live":
        date = args.date or datetime.now().strftime("%Y-%m-%d")
        asyncio.run(track_picks_live(date))

    print("\n" + "=" * 80)
    return 0


if __name__ == "__main__":
    sys.exit(main())
